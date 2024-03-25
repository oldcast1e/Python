'''
from calc.calculater import Calc

c = Calc(0)

c.add(100)
c.sub(50)
c.mult(2)
c.div(5)

res = c.getResult()

print('result = ', res)
'''


'''
from turtle import *
color('green', 'skyblue')
begin_fill()
pensize(5)
while True:
    forward(300)
    left(150)
    if abs(pos()) < 1:
        break
end_fill()
done()
'''


'''
import matplotlib.pyplot as ppl

x_nums = [1, 2, 3, 4, 7, 8]
y_nums = [3, 4, 6, 5, 4, 7]
ppl.plot(x_nums, y_nums)
ppl.show()
'''

'''
import openpyxl

work_book = openpyxl.Workbook()

work_sheet = work_book.create_sheet(index=0, title='Python Ch09 Exercise')
work_sheet['B2'] = 'Hello python world'

work_book.save(filename='ch09ex.xlsx')
'''


import openpyxl

wb = openpyxl.load_workbook('ch09ex.xlsx')
sheet_names = wb.sheetnames
print(sheet_names)

work_sheet = wb['Python Ch09 Exercise']

max_col = work_sheet.max_column
print('max column : ', max_col)
max_row = work_sheet.max_row
print('max row : ', max_row)

print('B2 셀의 값 : ', work_sheet['B2'].value)










